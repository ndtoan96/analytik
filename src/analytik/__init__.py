from TikTokApi import TikTokApi
import asyncio
import os
from pathlib import Path
import subprocess
import logging
from openpyxl import Workbook

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger("analytik")

BATCH_SIZE = 10
CSV_DIALECT = "excel"
URL_COL = 1
VIEW_COL = 2
LIKE_COL = 3
COMMENT_COL = 4
SHARE_COL = 5
SAVE_COL = 6


async def run():
    appdata_path = os.environ.get("APPDATA", None)
    if appdata_path is None:
        raise Exception("Cannot find APPDATA environment variable")
    analytik_path = Path(appdata_path).joinpath("analytik")
    if not analytik_path.exists():
        analytik_path.mkdir(parents=True)
    input_file_path = analytik_path.joinpath("links.txt")
    output_file_path = analytik_path.joinpath("report.xlsx")
    if not input_file_path.exists():
        with open(input_file_path, "w", encoding="utf-8") as f:
            f.write("Dán các link vào đây, lưu và đóng lại")
    subprocess.run(["notepad", input_file_path])
    with open(input_file_path, "r") as f:
        content = f.read()
    urls = [
        line.strip() for line in content.split("\n") if line.strip().startswith("http")
    ]
    wb = Workbook()
    sheet = wb.active
    sheet.cell(1, URL_COL).value = "Link"
    sheet.cell(1, VIEW_COL).value = "View"
    sheet.cell(1, LIKE_COL).value = "Like"
    sheet.cell(1, COMMENT_COL).value = "Comment"
    sheet.cell(1, SHARE_COL).value = "Share"
    sheet.cell(1, SAVE_COL).value = "Save"
    wb.save(output_file_path)
    async with TikTokApi() as api:
        await api.create_sessions(num_sessions=1, sleep_after=3, browser="chromium")
        row = 2
        for batch_num in range(len(urls) // BATCH_SIZE + 1):
            tasks = []
            async with asyncio.TaskGroup() as tg:
                for url in urls[
                    BATCH_SIZE * batch_num : min(
                        len(urls), BATCH_SIZE * (batch_num + 1)
                    )
                ]:
                    tasks.append((url, tg.create_task(api.video(url=url).info())))
            for url, task in tasks:
                stats = task.result().get("stats")
                like_count = stats.get("diggCount")
                comment_count = stats.get("commentCount")
                view_count = stats.get("playCount")
                share_count = stats.get("shareCount")
                save_count = int(stats.get("collectCount"))
                print(type(save_count))
                logger.info(
                    f"url: {url}, view: {view_count}, like: {like_count}, comment: {comment_count}, share: {share_count}, save: {save_count}"
                )
                sheet.cell(row, URL_COL).value = url
                sheet.cell(row, VIEW_COL).value = view_count
                sheet.cell(row, LIKE_COL).value = like_count
                sheet.cell(row, COMMENT_COL).value = comment_count
                sheet.cell(row, SHARE_COL).value = share_count
                sheet.cell(row, SAVE_COL).value = save_count
                wb.save(output_file_path)
                row += 1
    wb.close()
    subprocess.run(["cmd", "/C", "start", output_file_path])

def main():
    asyncio.run(run())
